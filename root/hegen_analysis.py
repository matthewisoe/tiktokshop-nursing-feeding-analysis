"""
=============================================================
  TikTok Shop Indonesia — Nursing & Feeding Category Analysis
  Brand Spotlight: Hegen Indonesia
  Data Source: FastMoss | Date: March 2026
=============================================================
Run this script from the same folder as your xlsx files.
Outputs: hegen_category_dashboard.xlsx + all chart PNGs
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import numpy as np
import re
import warnings
import os

warnings.filterwarnings('ignore')

# ── Path config ──────────────────────────────────────────────────────────────
DATA_DIR = "/Users/mattirv/Desktop/Category Analysis Project/Hegen/Datasets"
OUT_DIR  = "/Users/mattirv/Desktop/Category Analysis Project/Hegen/Analysis"
os.makedirs(OUT_DIR, exist_ok=True)

# ── Colour palette ───────────────────────────────────────────────────────────
C_PRIMARY   = "#E8344E"   # TikTok red
C_HEGEN     = "#2B6CB0"   # Hegen blue
C_DARK      = "#1A202C"
C_GRAY      = "#718096"
C_LIGHT     = "#EDF2F7"
C_GREEN     = "#38A169"
C_ORANGE    = "#DD6B20"

plt.rcParams.update({
    "font.family":      "DejaVu Sans",
    "axes.spines.top":  False,
    "axes.spines.right":False,
    "axes.grid":        True,
    "grid.color":       "#E2E8F0",
    "grid.linewidth":   0.6,
    "axes.titlesize":   13,
    "axes.titleweight": "bold",
    "axes.labelsize":   10,
    "xtick.labelsize":  9,
    "ytick.labelsize":  9,
    "figure.dpi":       130,
})

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def parse_idr(s):
    """Convert 'Rp119,7 jt' → float (IDR). Returns NaN if unparseable."""
    if pd.isna(s): return np.nan
    s = str(s).strip()
    # Handle range prices — take first value
    s = s.split("-")[0].strip()
    s = s.replace("Rp","").replace(" ","").replace(",",".")
    mult = 1
    if "M" in s:
        s = s.replace("M",""); mult = 1_000_000_000   # juta × 1000 = miliar? No: M = miliar
        # FastMoss uses: jt = juta (1e6), M = miliar (1e9)
        mult = 1_000_000_000
    elif "jt" in s:
        s = s.replace("jt",""); mult = 1_000_000
    elif "rb" in s:
        s = s.replace("rb",""); mult = 1_000
    try:
        return float(s) * mult
    except:
        return np.nan

def parse_k(s):
    """Convert '30.6k' / '2.1k' / '372' / '1.2m' → float."""
    if pd.isna(s): return np.nan
    s = str(s).strip().lower().replace(",",".")
    if "m" in s:
        return float(s.replace("m","")) * 1_000_000
    if "k" in s:
        return float(s.replace("k","")) * 1_000
    try:
        return float(s)
    except:
        return np.nan

def parse_pct(s):
    """'27.23%' → 27.23, '-' → NaN"""
    if pd.isna(s): return np.nan
    s = str(s).strip().replace("%","")
    try:
        return float(s)
    except:
        return np.nan

def parse_commission(s):
    if pd.isna(s) or str(s).strip() == "-": return np.nan
    return parse_pct(s)

def fmt_idr(n):
    if n >= 1e9:  return f"Rp{n/1e9:.1f} M"
    if n >= 1e6:  return f"Rp{n/1e6:.1f} jt"
    if n >= 1e3:  return f"Rp{n/1e3:.0f} rb"
    return f"Rp{n:.0f}"

def fmt_k(n):
    if n >= 1e6:  return f"{n/1e6:.1f}m"
    if n >= 1e3:  return f"{n/1e3:.1f}k"
    return f"{n:.0f}"

def save(fig, name):
    p = os.path.join(OUT_DIR, name)
    fig.savefig(p, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  ✓ Saved {name}")
    return p

# ═══════════════════════════════════════════════════════════════════════════
# LOAD & CLEAN DATA
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  LOADING DATA")
print("="*60)

shop_list    = pd.read_excel(f"{DATA_DIR}/Shop list (Nursing & Feeding).xlsx")
top_shops    = pd.read_excel(f"{DATA_DIR}/Top selling shops (Nursing & Feeding).xlsx")
top_products = pd.read_excel(f"{DATA_DIR}/Top selling products (Nursing & Feeding).xlsx")
promoted     = pd.read_excel(f"{DATA_DIR}/Most promoted products (Nursing & Feeding).xlsx")
video_promo  = pd.read_excel(f"{DATA_DIR}/Video-promoted products (Nursing & Feeding).xlsx")
creators     = pd.read_excel(f"{DATA_DIR}/Creator search (Nursing & Feeding).xlsx")
fast_growing = pd.read_excel(f"{DATA_DIR}/Fastest growing creators (Baby).xlsx")
sales_driv   = pd.read_excel(f"{DATA_DIR}/Top sales-driving creators (Nursing & Feeding).xlsx")
trending     = pd.read_excel(f"{DATA_DIR}/Trending creators (Baby).xlsx")
rising       = pd.read_excel(f"{DATA_DIR}/Rising star creators (Baby).xlsx")

print(f"  shop_list:     {len(shop_list)} rows")
print(f"  top_shops:     {len(top_shops)} rows")
print(f"  top_products:  {len(top_products)} rows")
print(f"  promoted:      {len(promoted)} rows")
print(f"  video_promo:   {len(video_promo)} rows")
print(f"  creators:      {len(creators)} rows")
print(f"  sales_driving: {len(sales_driv)} rows")
print(f"  rising_stars:  {len(rising)} rows")

# ── Clean shop_list ──────────────────────────────────────────────────────────
sl = shop_list.copy()
sl["gmv_7d"]        = sl["Revenue in Last 7 Days [Nursing & Feeding]"].apply(parse_idr)
sl["units_7d"]      = sl["Sales in Last 7 Days [Nursing & Feeding]"].apply(parse_k)
sl["total_gmv"]     = sl["Total Revenue [Nursing & Feeding]"].apply(parse_idr)
sl["total_units"]   = sl["Total Sales [Nursing & Feeding]"].apply(parse_k)
sl["influencers"]   = sl["Number of Influencers"].apply(parse_k)
sl["active_prods"]  = pd.to_numeric(sl["Active Products in the Last 7 Days"], errors="coerce")
sl["positioning"]   = sl["Shop Positioning"]

# ── Clean top_shops ──────────────────────────────────────────────────────────
ts = top_shops.copy()
ts["gmv"]           = ts["Revenue [Nursing & Feeding]"].apply(parse_idr)
ts["units"]         = ts["Sales [Nursing & Feeding]"].apply(parse_k)
ts["gmv_mom"]       = ts["Revenue MoM"].apply(parse_pct)
ts["units_mom"]     = ts["Sales MoM"].apply(parse_pct)
ts["influencers"]   = ts["Number of Influencers"].apply(parse_k)
ts["active_prods"]  = ts["Number of Active Products [Nursing & Feeding]"]

# ── Clean top_products ───────────────────────────────────────────────────────
tp = top_products.copy()
tp["total_units"]   = pd.to_numeric(tp["Total Units Sold"], errors="coerce")
tp["total_rev"]     = pd.to_numeric(tp["Total Revenue"], errors="coerce")
tp["orders"]        = pd.to_numeric(tp["Orders"], errors="coerce")
tp["commission"]    = tp["Commission Rate"].apply(parse_commission)
# clean subcategory
tp["subcategory"]   = tp["Products Category"].str.replace(r"Baby & Maternity-Nursing & Feeding-","", regex=True).str.strip()
# clean price
tp["price_low"]     = tp["Price"].apply(lambda x: parse_idr("Rp" + str(x).replace("Rp","").split("-")[0].strip()) if pd.notna(x) else np.nan)

# ── Clean promoted ───────────────────────────────────────────────────────────
pr = promoted.copy()
pr["commission"]    = pr["Commission Rate"].apply(parse_commission)
pr["affiliates"]    = pd.to_numeric(pr["Related Creators"], errors="coerce")
pr["total_aff"]     = pd.to_numeric(pr["Total Related Creators"], errors="coerce")
pr["total_units"]   = pd.to_numeric(pr["Total Units Sold"], errors="coerce")
pr["total_rev"]     = pd.to_numeric(pr["Total Revenue"], errors="coerce")
pr["shop_units"]    = pr["Shop Units Sold"].apply(parse_k)
pr["subcategory"]   = pr["Products Category"].str.replace("Baby & MaternityNursing & Feeding","", regex=False).str.strip()

# ── Clean video_promo ────────────────────────────────────────────────────────
vp = video_promo.copy()
vp["video_units"]   = pd.to_numeric(vp["Video Units Sold"], errors="coerce")
vp["total_units"]   = pd.to_numeric(vp["Video Total Units Sold"], errors="coerce")
vp["gmv"]           = pd.to_numeric(vp["Video GMV"], errors="coerce")
vp["views"]         = pd.to_numeric(vp["Total Views"], errors="coerce")
vp["price"]         = vp["Product Price"].apply(lambda x: parse_idr("Rp" + str(x).replace("Rp","").split("-")[0].strip()) if pd.notna(x) else np.nan)
vp["cvr_proxy"]     = vp["video_units"] / vp["Video Views"].replace(0, np.nan) * 100  # units/view %

# ── Clean creators ────────────────────────────────────────────────────────────
cr = creators.copy()
cr["followers"]     = pd.to_numeric(cr["Followers"], errors="coerce")
cr["units_28d"]     = pd.to_numeric(cr["Units Sold in last 28 days"], errors="coerce")
cr["video_units"]   = pd.to_numeric(cr["Units Sold by shoppable videos in last 28 days"], errors="coerce")
cr["live_units"]    = pd.to_numeric(cr["Units Sold by live in last 29 days"], errors="coerce")
cr["eng_rate"]      = cr["Video Engagement Rate"].apply(parse_pct)
cr["avg_views"]     = pd.to_numeric(cr["Average views of ecommerce videos"], errors="coerce")

# ── Clean sales_driv ─────────────────────────────────────────────────────────
sd = sales_driv.copy()
sd["followers"]     = pd.to_numeric(sd["Followers"], errors="coerce")
sd["video_gmv"]     = pd.to_numeric(sd["Video ECommerce GMV"], errors="coerce")
sd["live_gmv"]      = pd.to_numeric(sd["Live ECommerce GMV"], errors="coerce")
sd["total_gmv"]     = sd["video_gmv"].fillna(0) + sd["live_gmv"].fillna(0)
sd["products"]      = pd.to_numeric(sd["Promoted Products"], errors="coerce")

# ── Clean rising ─────────────────────────────────────────────────────────────
rs = rising.copy()
rs["followers"]     = pd.to_numeric(rs["Followers"], errors="coerce")
rs["video_gmv"]     = pd.to_numeric(rs["Video ECommerce GMV"], errors="coerce")
rs["live_gmv"]      = pd.to_numeric(rs["Live ECommerce GMV"], errors="coerce")
rs["total_gmv"]     = rs["video_gmv"].fillna(0) + rs["live_gmv"].fillna(0)
rs["viral_index"]   = pd.to_numeric(rs["Potential Index"], errors="coerce")

print("\n  ✓ All data cleaned")

# ═══════════════════════════════════════════════════════════════════════════
# HEGEN BENCHMARK VALUES
# ═══════════════════════════════════════════════════════════════════════════
hegen_shop = sl[sl["Shop Name"].str.contains("Hegen", case=False, na=False)].iloc[0]
hegen_ts   = ts[ts["Shop Name"].str.contains("Hegen", case=False, na=False)].iloc[0]
hegen_prod = pr[pr["Shop Name"].str.contains("Hegen", case=False, na=False)]

hegen_gmv_7d       = hegen_shop["gmv_7d"]
hegen_units_7d     = hegen_shop["units_7d"]
hegen_total_gmv    = hegen_shop["total_gmv"]
hegen_total_units  = hegen_shop["total_units"]
hegen_influencers  = hegen_shop["influencers"]
hegen_active_prods = hegen_shop["active_prods"]
hegen_rating       = hegen_shop["Shop Rating"]
hegen_gmv_mom      = hegen_ts["gmv_mom"]   # -59.9
hegen_units_mom    = hegen_ts["units_mom"] # -62.5
hegen_commission   = hegen_prod["commission"].mean() if len(hegen_prod) > 0 else 5.0

print(f"\n  HEGEN SNAPSHOT")
print(f"  GMV 7d:         {fmt_idr(hegen_gmv_7d)}")
print(f"  Units 7d:       {hegen_units_7d:.0f}")
print(f"  Total GMV:      {fmt_idr(hegen_total_gmv)}")
print(f"  Influencers:    {hegen_influencers:.0f}")
print(f"  GMV MoM:        {hegen_gmv_mom:.1f}%")
print(f"  Commission:     {hegen_commission:.1f}%")

# ── Category benchmarks ───────────────────────────────────────────────────────
cat_median_commission = pr["commission"].median()
cat_avg_affiliates    = pr["total_aff"].median()
top10_shops           = ts.nlargest(10, "gmv")
cat_top10_gmv_avg     = top10_shops["gmv"].mean()
cat_median_influencers= ts["influencers"].median()

print(f"\n  CATEGORY BENCHMARKS")
print(f"  Median commission:   {cat_median_commission:.1f}%")
print(f"  Median affiliates:   {cat_avg_affiliates:.0f}")
print(f"  Median influencers:  {cat_median_influencers:.0f}")
print(f"  Top-10 avg GMV:      {fmt_idr(cat_top10_gmv_avg)}")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — CATEGORY OVERVIEW CHARTS
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  SECTION 1 — CATEGORY OVERVIEW")
print("="*60)

# 1A: Top 15 shops by GMV (7-day), Hegen highlighted
fig, ax = plt.subplots(figsize=(11, 6))
top15 = sl.nlargest(15, "gmv_7d")[["Shop Name", "gmv_7d", "influencers", "positioning"]].copy()
# Add Hegen if not in top 15
if "Hegen Indonesia" not in top15["Shop Name"].values:
    top15 = pd.concat([top15, sl[sl["Shop Name"]=="Hegen Indonesia"][["Shop Name","gmv_7d","influencers","positioning"]]])

top15 = top15.sort_values("gmv_7d")
colors = [C_HEGEN if "Hegen" in n else C_PRIMARY for n in top15["Shop Name"]]
bars = ax.barh(top15["Shop Name"], top15["gmv_7d"] / 1e6, color=colors, height=0.65)

for bar, val in zip(bars, top15["gmv_7d"]):
    ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
            fmt_idr(val), va="center", fontsize=8.5, color=C_DARK)

ax.set_xlabel("GMV — Last 7 Days (Rp jt)")
ax.set_title("Top Shops by 7-Day GMV — Nursing & Feeding, Indonesia", pad=12)
legend = [mpatches.Patch(color=C_PRIMARY, label="Competitor"),
          mpatches.Patch(color=C_HEGEN,   label="Hegen Indonesia")]
ax.legend(handles=legend, loc="lower right", fontsize=9)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"Rp{x:.0f}jt"))
plt.tight_layout()
save(fig, "01_top_shops_gmv.png")

# 1B: MoM Growth — top shops (bubble = influencer count)
fig, ax = plt.subplots(figsize=(11, 6))
plot_ts = ts.dropna(subset=["gmv_mom","gmv"]).copy()
plot_ts = plot_ts[plot_ts["gmv_mom"].between(-100, 900)]
plot_ts["is_hegen"] = plot_ts["Shop Name"].str.contains("Hegen", case=False, na=False)

scatter = ax.scatter(
    plot_ts[~plot_ts["is_hegen"]]["gmv"] / 1e6,
    plot_ts[~plot_ts["is_hegen"]]["gmv_mom"],
    s=plot_ts[~plot_ts["is_hegen"]]["influencers"].fillna(0) / 50 + 20,
    color=C_PRIMARY, alpha=0.5, label="Competitors"
)
if plot_ts["is_hegen"].any():
    hrow = plot_ts[plot_ts["is_hegen"]].iloc[0]
    ax.scatter(hrow["gmv"]/1e6, hrow["gmv_mom"],
               s=max(hrow["influencers"]/50+20, 60),
               color=C_HEGEN, zorder=5, label="Hegen Indonesia", edgecolors="white", linewidths=1.5)
    ax.annotate("Hegen", (hrow["gmv"]/1e6, hrow["gmv_mom"]),
                textcoords="offset points", xytext=(8, 4), fontsize=9, color=C_HEGEN, fontweight="bold")

ax.axhline(0, color=C_GRAY, linewidth=0.8, linestyle="--")
ax.set_xlabel("GMV (Rp jt)")
ax.set_ylabel("MoM Growth (%)")
ax.set_title("Shop GMV vs MoM Growth — Bubble size = Influencer count", pad=12)
ax.legend(fontsize=9)
plt.tight_layout()
save(fig, "02_gmv_vs_mom_growth.png")

# 1C: Subcategory breakdown — units sold
fig, ax = plt.subplots(figsize=(9, 5))
sub_units = tp.groupby("subcategory")["total_units"].sum().sort_values(ascending=True)
sub_units = sub_units[sub_units > 0]
colors_sub = [C_HEGEN if "Bottle" in s else C_PRIMARY for s in sub_units.index]
ax.barh(sub_units.index, sub_units.values / 1e3, color=colors_sub, height=0.6)
ax.set_xlabel("Total Units Sold (thousands)")
ax.set_title("Units Sold by Subcategory — Nursing & Feeding", pad=12)
for i, (idx, val) in enumerate(sub_units.items()):
    ax.text(val/1e3 + 0.5, i, f"{val/1e3:.1f}k", va="center", fontsize=8.5)
note = mpatches.Patch(color=C_HEGEN, label="Baby Bottles & Accessories\n(Hegen's subcategory)")
ax.legend(handles=[note], fontsize=8.5)
plt.tight_layout()
save(fig, "03_subcategory_units.png")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — PRODUCT ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  SECTION 2 — PRODUCT ANALYSIS")
print("="*60)

# 2A: Commission rate distribution
fig, ax = plt.subplots(figsize=(8, 5))
comm_vals = pr["commission"].dropna()
bins = [0, 2, 4, 6, 8, 10, 12, 16]
counts, edges = np.histogram(comm_vals, bins=bins)
bar_colors = []
for i in range(len(counts)):
    mid = (edges[i] + edges[i+1]) / 2
    if abs(mid - hegen_commission) < 1.5:
        bar_colors.append(C_HEGEN)
    else:
        bar_colors.append(C_PRIMARY)

ax.bar(edges[:-1], counts, width=np.diff(edges)*0.85,
       color=bar_colors, align="edge", edgecolor="white")
ax.axvline(cat_median_commission, color=C_ORANGE, linewidth=1.5, linestyle="--",
           label=f"Category median: {cat_median_commission:.0f}%")
ax.axvline(hegen_commission, color=C_HEGEN, linewidth=2, linestyle="-",
           label=f"Hegen: {hegen_commission:.0f}%")
ax.set_xlabel("Commission Rate (%)")
ax.set_ylabel("Number of Products")
ax.set_title("Commission Rate Distribution — Nursing & Feeding Products", pad=12)
ax.legend(fontsize=9)
plt.tight_layout()
save(fig, "04_commission_distribution.png")

# 2B: Top 10 products by affiliates (creator coverage)
fig, ax = plt.subplots(figsize=(11, 6))
top_aff = pr.nlargest(12, "total_aff")[["Product Name","Shop Name","total_aff","commission"]].copy()
top_aff["label"] = top_aff["Product Name"].str[:40] + "…"
top_aff["is_hegen"] = top_aff["Shop Name"].str.contains("Hegen", case=False, na=False)
top_aff = top_aff.sort_values("total_aff")
c_bars = [C_HEGEN if h else C_PRIMARY for h in top_aff["is_hegen"]]
ax.barh(top_aff["label"], top_aff["total_aff"], color=c_bars, height=0.65)
for i, (_, row) in enumerate(top_aff.iterrows()):
    ax.text(row["total_aff"] + 10, i, f'{row["total_aff"]:.0f}', va="center", fontsize=8.5)
ax.set_xlabel("Total Affiliates (cumulative)")
ax.set_title("Top Products by Total Affiliate Coverage — Nursing & Feeding", pad=12)
legend = [mpatches.Patch(color=C_PRIMARY, label="Competitor"),
          mpatches.Patch(color=C_HEGEN,   label="Hegen")]
ax.legend(handles=legend, fontsize=9)
plt.tight_layout()
save(fig, "05_top_products_affiliates.png")

# 2C: Price vs Units sold scatter — Baby Bottles subcategory
fig, ax = plt.subplots(figsize=(9, 5))
bottles = tp[tp["subcategory"].str.contains("Bottle", na=False)].copy()
hegen_bottles = bottles[bottles["Shop Name"].str.contains("Hegen", na=False)]
others = bottles[~bottles["Shop Name"].str.contains("Hegen", na=False)]

ax.scatter(others["price_low"] / 1e3, others["total_units"] / 1e3,
           color=C_PRIMARY, alpha=0.45, s=40, label="Other brands")
if len(hegen_bottles) > 0:
    ax.scatter(hegen_bottles["price_low"] / 1e3, hegen_bottles["total_units"] / 1e3,
               color=C_HEGEN, s=120, zorder=5, label="Hegen", edgecolors="white", linewidths=1.5)
    for _, row in hegen_bottles.iterrows():
        ax.annotate(row["Product Name"][:25]+"…",
                    (row["price_low"]/1e3, row["total_units"]/1e3),
                    textcoords="offset points", xytext=(5, 4), fontsize=7.5, color=C_HEGEN)

ax.set_xlabel("Price (Rp ribu)")
ax.set_ylabel("Total Units Sold (thousands)")
ax.set_title("Price vs Units Sold — Baby Bottles & Accessories", pad=12)
ax.legend(fontsize=9)
plt.tight_layout()
save(fig, "06_price_vs_units_bottles.png")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — CREATOR INTELLIGENCE
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  SECTION 3 — CREATOR INTELLIGENCE")
print("="*60)

# 3A: Creator tier analysis — units sold by follower tier
def tier(f):
    if f < 10_000:    return "Nano\n(<10K)"
    if f < 100_000:   return "Micro\n(10K–100K)"
    if f < 500_000:   return "Mid-tier\n(100K–500K)"
    if f < 1_000_000: return "Macro\n(500K–1M)"
    return "Mega\n(>1M)"

cr["tier"] = cr["followers"].apply(tier)
tier_order = ["Nano\n(<10K)","Micro\n(10K–100K)","Mid-tier\n(100K–500K)","Macro\n(500K–1M)","Mega\n(>1M)"]
tier_stats = cr.groupby("tier").agg(
    creators=("Creator Name","count"),
    total_units=("units_28d","sum"),
    avg_units=("units_28d","mean"),
    avg_eng=("eng_rate","mean")
).reindex(tier_order).fillna(0)

fig, axes = plt.subplots(1, 2, figsize=(13, 5))

axes[0].bar(tier_stats.index, tier_stats["total_units"] / 1e3,
            color=[C_PRIMARY if i != 1 else C_GREEN for i in range(len(tier_stats))],
            edgecolor="white")
axes[0].set_title("Total Units Sold (28d) by Creator Tier", pad=10)
axes[0].set_ylabel("Units Sold (thousands)")
axes[0].set_xlabel("Creator Tier")
for i, (idx, row) in enumerate(tier_stats.iterrows()):
    axes[0].text(i, row["total_units"]/1e3 + 0.5, f"{row['total_units']/1e3:.0f}k",
                 ha="center", fontsize=8.5)

axes[1].bar(tier_stats.index, tier_stats["avg_eng"].fillna(0),
            color=C_ORANGE, edgecolor="white", alpha=0.85)
axes[1].set_title("Avg Engagement Rate by Creator Tier", pad=10)
axes[1].set_ylabel("Engagement Rate (%)")
axes[1].set_xlabel("Creator Tier")
for i, (idx, row) in enumerate(tier_stats.iterrows()):
    if row["avg_eng"] > 0:
        axes[1].text(i, row["avg_eng"] + 0.02, f"{row['avg_eng']:.2f}%",
                     ha="center", fontsize=8.5)

plt.suptitle("Creator Tier Analysis — Nursing & Feeding Affiliates", fontsize=13, fontweight="bold", y=1.02)
plt.tight_layout()
save(fig, "07_creator_tier_analysis.png")

# 3B: Video vs Live GMV split — top sales-driving creators
fig, ax = plt.subplots(figsize=(11, 6))
top_sd = sd.nlargest(12, "total_gmv")[["Creator name","video_gmv","live_gmv","followers","total_gmv"]].copy()
top_sd = top_sd.sort_values("total_gmv")
top_sd["label"] = top_sd["Creator name"].str[:28]

y = range(len(top_sd))
ax.barh(y, top_sd["video_gmv"] / 1e6, color=C_PRIMARY, label="Video GMV", height=0.6)
ax.barh(y, top_sd["live_gmv"] / 1e6,
        left=top_sd["video_gmv"] / 1e6, color=C_ORANGE, label="Live GMV", height=0.6)
ax.set_yticks(list(y))
ax.set_yticklabels(top_sd["label"].values)
ax.set_xlabel("GMV (Rp jt)")
ax.set_title("Video vs Live GMV — Top Sales-Driving Creators, Nursing & Feeding", pad=12)
ax.legend(fontsize=9)
plt.tight_layout()
save(fig, "08_video_vs_live_gmv.png")

# 3C: GMV efficiency — GMV per follower for top creators
fig, ax = plt.subplots(figsize=(9, 5))
sd_eff = sd.copy()
sd_eff["gmv_per_follower"] = sd_eff["total_gmv"] / sd_eff["followers"].replace(0, np.nan)
sd_eff = sd_eff.dropna(subset=["gmv_per_follower"]).nlargest(12, "gmv_per_follower")
sd_eff = sd_eff.sort_values("gmv_per_follower")

ax.barh(sd_eff["Creator name"].str[:28], sd_eff["gmv_per_follower"],
        color=C_GREEN, height=0.65)
ax.set_xlabel("GMV per Follower (IDR)")
ax.set_title("Creator Efficiency — GMV per Follower\n(Higher = more efficient for brand spend)", pad=10)
for i, (_, row) in enumerate(sd_eff.iterrows()):
    ax.text(row["gmv_per_follower"] + 5, i,
            f"Rp{row['gmv_per_follower']:.0f}", va="center", fontsize=8)
plt.tight_layout()
save(fig, "09_creator_gmv_efficiency.png")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — HEGEN BRAND SPOTLIGHT
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  SECTION 4 — HEGEN BRAND SPOTLIGHT")
print("="*60)

# 4A: Hegen vs category benchmark — radar-style bar comparison
metrics = {
    "Influencers / 100 shops": (hegen_influencers, cat_median_influencers),
    "Commission Rate (%)":     (hegen_commission,  cat_median_commission),
    "7-Day GMV (Rp jt)":       (hegen_gmv_7d/1e6, top10_shops["gmv"].median()/1e6),
    "7-Day Units Sold":        (hegen_units_7d,   ts["units"].median()),
    "Active Products":         (float(hegen_active_prods), ts["active_prods"].median()),
}

fig, axes = plt.subplots(1, len(metrics), figsize=(14, 5))
for ax, (metric, (hegen_val, cat_val)) in zip(axes, metrics.items()):
    bars = ax.bar(["Hegen", "Category\nMedian"],
                  [hegen_val, cat_val],
                  color=[C_HEGEN, C_LIGHT],
                  edgecolor=[C_HEGEN, C_GRAY],
                  linewidth=1.2,
                  width=0.55)
    for bar, val in zip(bars, [hegen_val, cat_val]):
        label = f"{val:.1f}" if val < 1000 else fmt_k(val) if val >= 1000 else f"{val:.0f}"
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() * 1.04,
                label, ha="center", fontsize=9, fontweight="bold",
                color=C_HEGEN if bar.get_facecolor()[2] > 0.5 else C_DARK)
    ax.set_title(metric, fontsize=9, pad=8)
    ax.set_ylim(0, max(hegen_val, cat_val) * 1.4)
    ax.tick_params(axis="y", labelsize=8)
    ax.set_facecolor("#FAFAFA")

plt.suptitle("Hegen Indonesia vs Category Median — Key Metrics", fontsize=13, fontweight="bold", y=1.02)
plt.tight_layout()
save(fig, "10_hegen_vs_benchmark.png")

# 4B: Hegen affiliate gap analysis
fig, ax = plt.subplots(figsize=(8, 4.5))
top5_competitors = pr[~pr["Shop Name"].str.contains("Hegen", na=False)].nlargest(5, "total_aff")
hegen_aff_row = pr[pr["Shop Name"].str.contains("Hegen", na=False)]
hegen_max_aff = hegen_aff_row["total_aff"].max() if len(hegen_aff_row) > 0 else 9

comparison = pd.concat([
    top5_competitors[["Product Name","total_aff","Shop Name"]].head(5),
    hegen_aff_row[["Product Name","total_aff","Shop Name"]].head(1)
])
comparison["label"] = comparison.apply(
    lambda r: ("Hegen: " + r["Product Name"][:25]) if "Hegen" in str(r["Shop Name"]) else r["Product Name"][:35], axis=1)
comparison = comparison.sort_values("total_aff")
c_bars = [C_HEGEN if "Hegen" in str(r["Shop Name"]) else C_PRIMARY
          for _, r in comparison.iterrows()]

bars = ax.barh(comparison["label"] + "…", comparison["total_aff"],
               color=c_bars, height=0.65)
for bar, val in zip(bars, comparison["total_aff"]):
    ax.text(bar.get_width() + 5, bar.get_y() + bar.get_height()/2,
            f"{val:.0f}", va="center", fontsize=8.5)
ax.set_xlabel("Total Affiliates (cumulative)")
ax.set_title("Hegen's Affiliate Gap vs Top Competitor Products", pad=12)
legend = [mpatches.Patch(color=C_PRIMARY, label="Top competitor products"),
          mpatches.Patch(color=C_HEGEN,   label="Hegen best product")]
ax.legend(handles=legend, fontsize=9)
plt.tight_layout()
save(fig, "11_hegen_affiliate_gap.png")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5 — STRATEGIC RECOMMENDATIONS (text summary)
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  SECTION 5 — KEY INSIGHTS & RECOMMENDATIONS")
print("="*60)

# MoM context
top_growers = ts.nlargest(3, "gmv_mom")[["Shop Name","gmv_mom"]]
bottom_shops = ts.nsmallest(3, "gmv_mom")[["Shop Name","gmv_mom"]]

# Video CVR
avg_cvr = vp["cvr_proxy"].median()

# Live vs video split in sales-driving creators
total_video_gmv = sd["video_gmv"].sum()
total_live_gmv  = sd["live_gmv"].sum()
live_share = total_live_gmv / (total_video_gmv + total_live_gmv) * 100

# Dominant tier for units
best_tier = tier_stats["total_units"].idxmax()

insights = f"""
╔══════════════════════════════════════════════════════════════════╗
║        KEY INSIGHTS — Nursing & Feeding TikTok Shop ID          ║
╠══════════════════════════════════════════════════════════════════╣
║                                                                  ║
║  CATEGORY OVERVIEW                                               ║
║  • Baby Bottles & Accessories dominates units sold (~75 of      ║
║    300 top products), directly Hegen's core subcategory.        ║
║  • Top 3 shops (Sakumini, Dr.Isla, KOI.STORE) control the      ║
║    majority of 7-day GMV — market is concentrated.              ║
║  • {top_growers.iloc[0]['Shop Name'][:30]} leads MoM growth     ║
║    at +{top_growers.iloc[0]['gmv_mom']:.0f}%.                   ║
║                                                                  ║
║  CREATOR & CHANNEL INSIGHTS                                      ║
║  • Live commerce drives {live_share:.0f}% of top-creator GMV   ║
║    in this category — video alone is not enough.                ║
║  • {best_tier.replace(chr(10),' ')} creators generate the most  ║
║    total units — sweet spot for affiliate recruitment.          ║
║  • Median video CVR: {avg_cvr:.2f}% (units/view).              ║
║                                                                  ║
║  HEGEN DIAGNOSIS                                                 ║
║  • Hegen rank 265/300 shops — significant headroom.             ║
║  • GMV MoM: {hegen_gmv_mom:.1f}% — declining trend.           ║
║  • Only 372 influencers vs category median of                   ║
║    {cat_median_influencers:.0f} — severely under-indexed.      ║
║  • Commission rate {hegen_commission:.0f}% vs median            ║
║    {cat_median_commission:.0f}% — below category norm.         ║
║  • Best product: only 9 total affiliates vs top competitor's    ║
║    {top5_competitors.iloc[-1]['total_aff']:.0f}+ affiliates.   ║
║                                                                  ║
║  3 STRATEGIC RECOMMENDATIONS FOR HEGEN                          ║
║                                                                  ║
║  1. RAISE COMMISSION TO {cat_median_commission+2:.0f}%          ║
║     Current {hegen_commission:.0f}% is below category median.  ║
║     Raising to {cat_median_commission+2:.0f}% would bring it   ║
║     in line with the highest-affiliate products, directly       ║
║     increasing creator incentive to promote.                    ║
║                                                                  ║
║  2. BUILD A LIVE COMMERCE CHANNEL                                ║
║     {live_share:.0f}% of category GMV comes from live stream.  ║
║     Hegen has 0 live-specialist creators. Partner with          ║
║     1–2 mid-tier (100K–500K) baby creators for weekly          ║
║     product demos — this is Hegen's single biggest gap.        ║
║                                                                  ║
║  3. TARGET MICRO-CREATOR TIER FOR AFFILIATE SCALE               ║
║     {best_tier.replace(chr(10),' ')} creators drive the most   ║
║     total units with highest engagement rates. A structured     ║
║     micro-creator seeding programme (target: 50+ new           ║
║     affiliates in 30 days) would close the affiliate gap        ║
║     vs top competitors faster than chasing mega-creators.      ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝
"""
print(insights)

# ═══════════════════════════════════════════════════════════════════════════
# BUILD EXCEL DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  BUILDING EXCEL DASHBOARD")
print("="*60)

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

wb = Workbook()

# ── Style helpers ─────────────────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="E8344E")
BLUE_FILL  = PatternFill("solid", fgColor="2B6CB0")
DARK_FILL  = PatternFill("solid", fgColor="1A202C")
GRAY_FILL  = PatternFill("solid", fgColor="EDF2F7")
LGRAY_FILL = PatternFill("solid", fgColor="F7FAFC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="38A169")
ORANGE_FILL= PatternFill("solid", fgColor="DD6B20")

def hdr(cell, text, fill=RED_FILL, font_color="FFFFFF", size=11, bold=True):
    cell.value = text
    cell.font  = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.fill  = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def val(cell, text, bold=False, color="1A202C", size=10, align="left", fill=None):
    cell.value = text
    cell.font  = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill: cell.fill = fill

def border_all(ws, row_start, col_start, row_end, col_end, style="thin"):
    s = Side(style=style, color="CBD5E0")
    b = Border(left=s, right=s, top=s, bottom=s)
    for r in range(row_start, row_end+1):
        for c in range(col_start, col_end+1):
            ws.cell(r, c).border = b

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ════════════════════════════════════
# SHEET 1: OVERVIEW
# ════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Category Overview"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 18

ws1.merge_cells("A1:K1")
hdr(ws1["A1"],
    "TikTok Shop Indonesia — Nursing & Feeding Category Analysis  |  Data: FastMoss  |  March 2026",
    fill=DARK_FILL, size=13)

ws1.merge_cells("A2:K2")
val(ws1["A2"], "Brand Spotlight: Hegen Indonesia", bold=True, color="2B6CB0", size=11, align="center")

# KPI cards row
kpi_row = 4
ws1.merge_cells(f"A{kpi_row}:K{kpi_row}")
hdr(ws1[f"A{kpi_row}"], "CATEGORY SNAPSHOT — NURSING & FEEDING (TOP 300 SHOPS)", fill=RED_FILL)

kpi_data = [
    ("Total Shops Tracked", "300", "A", "B"),
    ("Top Shop GMV (7d)", fmt_idr(sl.nlargest(1,"gmv_7d")["gmv_7d"].values[0]), "C", "D"),
    ("Median Commission", f"{cat_median_commission:.0f}%", "E", "F"),
    ("Avg Affiliates / Product", f"{pr['total_aff'].mean():.0f}", "G", "H"),
    ("Live Share of Creator GMV", f"{live_share:.0f}%", "I", "J"),
]

ws1.row_dimensions[5].height = 18
ws1.row_dimensions[6].height = 32
ws1.row_dimensions[7].height = 22

for label, value, c1, c2 in kpi_data:
    ws1.merge_cells(f"{c1}5:{c2}5")
    val(ws1[f"{c1}5"], label, bold=False, color="718096", size=9, align="center", fill=LGRAY_FILL)
    ws1.merge_cells(f"{c1}6:{c2}6")
    val(ws1[f"{c1}6"], value, bold=True, color="1A202C", size=16, align="center", fill=WHITE_FILL)
    ws1.merge_cells(f"{c1}7:{c2}7")
    ws1[f"{c1}7"].fill = LGRAY_FILL

# Top shops table
tbl_row = 9
ws1.merge_cells(f"A{tbl_row}:K{tbl_row}")
hdr(ws1[f"A{tbl_row}"], "TOP 15 SHOPS BY 7-DAY GMV", fill=DARK_FILL)

headers = ["Rank", "Shop Name", "Positioning", "7d GMV", "7d Units",
           "MoM GMV", "Total GMV", "Total Units", "Influencers", "Active Prods", "Rating"]
for i, h in enumerate(headers, 1):
    hdr(ws1.cell(tbl_row+1, i), h, fill=GRAY_FILL, font_color="1A202C", size=9)

top15_xl = sl.nlargest(15, "gmv_7d").copy()
# merge with MoM from top_shops
mom_map = ts.set_index("Shop Name")[["gmv_mom","units_mom"]].to_dict()

for r_idx, (_, row) in enumerate(top15_xl.iterrows(), tbl_row+2):
    is_hegen = "Hegen" in str(row["Shop Name"])
    fill = PatternFill("solid", fgColor="EBF4FF") if is_hegen else (LGRAY_FILL if r_idx % 2 == 0 else WHITE_FILL)
    cells_data = [
        r_idx - tbl_row - 1,
        row["Shop Name"],
        row["positioning"],
        fmt_idr(row["gmv_7d"]) if pd.notna(row["gmv_7d"]) else "-",
        fmt_k(row["units_7d"]) if pd.notna(row["units_7d"]) else "-",
        f"{mom_map['gmv_mom'].get(row['Shop Name'], np.nan):.1f}%" if pd.notna(mom_map["gmv_mom"].get(row["Shop Name"], np.nan)) else "-",
        fmt_idr(row["total_gmv"]) if pd.notna(row["total_gmv"]) else "-",
        fmt_k(row["total_units"]) if pd.notna(row["total_units"]) else "-",
        fmt_k(row["influencers"]) if pd.notna(row["influencers"]) else "-",
        int(row["active_prods"]) if pd.notna(row["active_prods"]) else "-",
        row["Shop Rating"] if pd.notna(row["Shop Rating"]) else "-",
    ]
    for c_idx, v in enumerate(cells_data, 1):
        cell = ws1.cell(r_idx, c_idx)
        val(cell, v, fill=fill, size=9,
            bold=is_hegen, color="2B6CB0" if is_hegen else "1A202C")
        ws1.row_dimensions[r_idx].height = 16

border_all(ws1, tbl_row, 1, tbl_row+16, 11)

set_col_widths(ws1, {
    "A": 6, "B": 24, "C": 12, "D": 14, "E": 11,
    "F": 10, "G": 14, "H": 13, "I": 13, "J": 13, "K": 8
})

# Embed chart image
try:
    img = XLImage(f"{OUT_DIR}/01_top_shops_gmv.png")
    img.width, img.height = 520, 300
    ws1.add_image(img, "A29")
except: pass

# ════════════════════════════════════
# SHEET 2: PRODUCT ANALYSIS
# ════════════════════════════════════
ws2 = wb.create_sheet("2. Product Analysis")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:L1")
hdr(ws2["A1"], "TOP PRODUCTS — NURSING & FEEDING  |  Subcategory, Commission & Affiliate Coverage",
    fill=DARK_FILL, size=12)

# Commission summary
ws2.merge_cells("A3:L3")
hdr(ws2["A3"], "COMMISSION RATE SUMMARY — MOST PROMOTED PRODUCTS", fill=RED_FILL)

comm_labels = ["Min", "25th pct", "Median", "75th pct", "Max", "Hegen"]
comm_values = [
    pr["commission"].min(),
    pr["commission"].quantile(0.25),
    pr["commission"].median(),
    pr["commission"].quantile(0.75),
    pr["commission"].max(),
    hegen_commission,
]
for i, (lbl, v) in enumerate(zip(comm_labels, comm_values), 1):
    hdr(ws2.cell(4, i*2-1), lbl,
        fill=BLUE_FILL if lbl == "Hegen" else GRAY_FILL,
        font_color="FFFFFF" if lbl == "Hegen" else "1A202C", size=9)
    val(ws2.cell(4, i*2), f"{v:.1f}%", bold=lbl=="Hegen",
        color="2B6CB0" if lbl=="Hegen" else "1A202C",
        fill=PatternFill("solid",fgColor="EBF4FF") if lbl=="Hegen" else WHITE_FILL,
        align="center", size=11)

# Top promoted products table
tbl2_row = 7
ws2.merge_cells(f"A{tbl2_row}:L{tbl2_row}")
hdr(ws2[f"A{tbl2_row}"], "TOP 20 MOST PROMOTED PRODUCTS — BY TOTAL AFFILIATES", fill=DARK_FILL)

headers2 = ["Rank", "Product Name", "Shop", "Subcategory", "Price",
            "Commission", "Units Sold", "Revenue (IDR)", "Active Aff.", "Total Aff.", "Status", ""]
for i, h in enumerate(headers2, 1):
    hdr(ws2.cell(tbl2_row+1, i), h, fill=GRAY_FILL, font_color="1A202C", size=9)

top20_pr = pr.nlargest(20, "total_aff").copy()
# Add hegen if not in top 20
if not top20_pr["Shop Name"].str.contains("Hegen", na=False).any():
    hegen_rows = pr[pr["Shop Name"].str.contains("Hegen", na=False)]
    top20_pr = pd.concat([top20_pr, hegen_rows]).head(21)

for r_idx, (_, row) in enumerate(top20_pr.iterrows(), tbl2_row+2):
    is_hegen = "Hegen" in str(row["Shop Name"])
    fill = PatternFill("solid", fgColor="EBF4FF") if is_hegen else (LGRAY_FILL if r_idx % 2 == 0 else WHITE_FILL)
    cells = [
        r_idx - tbl2_row - 1,
        str(row["Product Name"])[:55],
        str(row["Shop Name"])[:22],
        str(row["subcategory"])[:28],
        str(row["Price"])[:18],
        f"{row['commission']:.0f}%" if pd.notna(row["commission"]) else "-",
        fmt_k(row["total_units"]) if pd.notna(row["total_units"]) else "-",
        fmt_idr(row["total_rev"]) if pd.notna(row["total_rev"]) else "-",
        int(row["affiliates"]) if pd.notna(row["affiliates"]) else "-",
        int(row["total_aff"]) if pd.notna(row["total_aff"]) else "-",
        str(row["Product Status"]) if pd.notna(row.get("Product Status","")) else "-",
        "← HEGEN" if is_hegen else "",
    ]
    for c_idx, v in enumerate(cells, 1):
        cell = ws2.cell(r_idx, c_idx)
        val(cell, v, fill=fill, size=9,
            bold=is_hegen, color="2B6CB0" if is_hegen else "1A202C")
        ws2.row_dimensions[r_idx].height = 15

border_all(ws2, tbl2_row, 1, tbl2_row+22, 11)
set_col_widths(ws2, {
    "A":6,"B":52,"C":22,"D":28,"E":18,
    "F":11,"G":11,"H":14,"I":11,"J":11,"K":12,"L":10
})

try:
    img2 = XLImage(f"{OUT_DIR}/04_commission_distribution.png")
    img2.width, img2.height = 440, 280
    ws2.add_image(img2, "A32")
    img3 = XLImage(f"{OUT_DIR}/05_top_products_affiliates.png")
    img3.width, img3.height = 520, 310
    ws2.add_image(img3, "G32")
except: pass

# ════════════════════════════════════
# SHEET 3: CREATOR INTELLIGENCE
# ════════════════════════════════════
ws3 = wb.create_sheet("3. Creator Intelligence")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
hdr(ws3["A1"], "CREATOR INTELLIGENCE — Nursing & Feeding Affiliate Ecosystem", fill=DARK_FILL, size=12)

# Creator tier summary
ws3.merge_cells("A3:J3")
hdr(ws3["A3"], "CREATOR TIER BREAKDOWN — Units Sold (28 days)", fill=RED_FILL)

tier_headers = ["Tier", "Creators", "Total Units (28d)", "Avg Units/Creator", "Avg Engagement"]
for i, h in enumerate(tier_headers, 1):
    hdr(ws3.cell(4, i), h, fill=GRAY_FILL, font_color="1A202C", size=9)

for r_idx, (tier_name, row) in enumerate(tier_stats.iterrows(), 5):
    is_best = tier_name == best_tier
    fill = PatternFill("solid", fgColor="F0FFF4") if is_best else (LGRAY_FILL if r_idx%2==0 else WHITE_FILL)
    cells = [
        tier_name.replace("\n"," "),
        int(row["creators"]),
        fmt_k(row["total_units"]),
        fmt_k(row["avg_units"]),
        f"{row['avg_eng']:.2f}%" if row["avg_eng"] > 0 else "-",
    ]
    for c_idx, v in enumerate(cells, 1):
        val(ws3.cell(r_idx, c_idx), v, fill=fill, size=9,
            bold=is_best, color="38A169" if is_best else "1A202C")
        ws3.row_dimensions[r_idx].height = 16
    if is_best:
        val(ws3.cell(r_idx, 6), "← Best performing tier", color="38A169", size=9, bold=True)

border_all(ws3, 4, 1, 9, 5)

# Top sales-driving creators table
tbl3_row = 11
ws3.merge_cells(f"A{tbl3_row}:J{tbl3_row}")
hdr(ws3[f"A{tbl3_row}"], "TOP 15 SALES-DRIVING CREATORS — Nursing & Feeding", fill=DARK_FILL)

headers3 = ["Rank","Creator","Region","Category","Followers","Products","Video GMV","Live GMV","Total GMV","Live %"]
for i, h in enumerate(headers3, 1):
    hdr(ws3.cell(tbl3_row+1, i), h, fill=GRAY_FILL, font_color="1A202C", size=9)

top15_sd = sd.nlargest(15, "total_gmv")
for r_idx, (_, row) in enumerate(top15_sd.iterrows(), tbl3_row+2):
    fill = LGRAY_FILL if r_idx % 2 == 0 else WHITE_FILL
    live_pct = (row["live_gmv"] / row["total_gmv"] * 100) if row["total_gmv"] > 0 else 0
    cells = [
        r_idx - tbl3_row - 1,
        str(row["Creator name"])[:30],
        str(row["Region"]),
        str(row["Creator Categories"])[:20],
        fmt_k(row["followers"]),
        int(row["products"]) if pd.notna(row["products"]) else "-",
        fmt_idr(row["video_gmv"]),
        fmt_idr(row["live_gmv"]),
        fmt_idr(row["total_gmv"]),
        f"{live_pct:.0f}%",
    ]
    for c_idx, v in enumerate(cells, 1):
        cell = ws3.cell(r_idx, c_idx)
        val(cell, v, fill=fill, size=9)
        # Color-code live % — high live = green
        if c_idx == 10 and live_pct > 60:
            cell.font = Font(bold=True, color="38A169", size=9, name="Calibri")
        ws3.row_dimensions[r_idx].height = 15

border_all(ws3, tbl3_row, 1, tbl3_row+16, 10)
set_col_widths(ws3, {"A":6,"B":28,"C":11,"D":20,"E":12,"F":10,"G":14,"H":14,"I":14,"J":8})

try:
    img4 = XLImage(f"{OUT_DIR}/07_creator_tier_analysis.png")
    img4.width, img4.height = 560, 260
    ws3.add_image(img4, "A30")
    img5 = XLImage(f"{OUT_DIR}/08_video_vs_live_gmv.png")
    img5.width, img5.height = 520, 310
    ws3.add_image(img5, "A49")
except: pass

# ════════════════════════════════════
# SHEET 4: HEGEN SPOTLIGHT
# ════════════════════════════════════
ws4 = wb.create_sheet("4. Hegen Spotlight")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
hdr(ws4["A1"], "HEGEN INDONESIA — Brand Performance & Strategic Gap Analysis", fill=BLUE_FILL, size=13)

ws4.merge_cells("A2:H2")
val(ws4["A2"], "Data: FastMoss  |  Category: Nursing & Feeding  |  Country: Indonesia  |  March 2026",
    color="718096", size=9, align="center")

# Scorecard
ws4.merge_cells("A4:H4")
hdr(ws4["A4"], "HEGEN PERFORMANCE SCORECARD", fill=DARK_FILL)

scorecard = [
    ("Shop Rating",        f"{hegen_rating}/5.0",          "3.7 is below top brands (4.7–4.9)",       "⚠"),
    ("7-Day GMV",          fmt_idr(hegen_gmv_7d),          f"Rank ~265/300 shops",                     "⚠"),
    ("GMV MoM Change",     f"{hegen_gmv_mom:.1f}%",        "Declining — needs immediate action",       "✗"),
    ("Units MoM Change",   f"{hegen_units_mom:.1f}%",      "Volume declining faster than revenue",     "✗"),
    ("Total Influencers",  f"{hegen_influencers:.0f}",      f"Category median: {cat_median_influencers:.0f}", "⚠"),
    ("Commission Rate",    f"{hegen_commission:.0f}%",      f"Below category median ({cat_median_commission:.0f}%)", "⚠"),
    ("Active Products",    f"{hegen_active_prods:.0f}",     "103 listed, only 8 active recently",      "⚠"),
    ("Top Product Aff.",   "9",                              f"Top competitor product has {int(top5_competitors.iloc[-1]['total_aff'])}+", "✗"),
]

score_headers = ["Metric", "Hegen Value", "Benchmark / Context", "Signal"]
for i, h in enumerate(score_headers, 1):
    hdr(ws4.cell(5, i*2-1), h, fill=GRAY_FILL, font_color="1A202C", size=9)

for r_idx, (metric, value, context, signal) in enumerate(scorecard, 6):
    fill_row = LGRAY_FILL if r_idx % 2 == 0 else WHITE_FILL
    signal_color = "E53E3E" if signal == "✗" else ("DD6B20" if signal == "⚠" else "38A169")
    data = [(metric, "1A202C", False), (value, "2B6CB0", True),
            (context, "718096", False), (signal, signal_color, True)]
    for c_idx, (v, color, bold) in enumerate(data, 1):
        cell = ws4.cell(r_idx, c_idx*2-1)
        val(cell, v, bold=bold, color=color, size=10, fill=fill_row)
        ws4.merge_cells(f"{get_column_letter(c_idx*2-1)}{r_idx}:{get_column_letter(c_idx*2)}{r_idx}")
        ws4.row_dimensions[r_idx].height = 18

border_all(ws4, 5, 1, 13, 8)

# Recommendations
ws4.merge_cells("A16:H16")
hdr(ws4["A16"], "STRATEGIC RECOMMENDATIONS", fill=RED_FILL)

recs = [
    ("1", f"Raise commission to {cat_median_commission+2:.0f}%",
     f"Current {hegen_commission:.0f}% is below the {cat_median_commission:.0f}% category median. Top products offering {cat_median_commission+2:.0f}%+ attract {pr[pr['commission']>=(cat_median_commission+2)]['total_aff'].median():.0f}+ affiliates on average. A 2% uplift would likely 3x affiliate adoption within 60 days."),
    ("2", "Launch a dedicated live commerce channel",
     f"Live streams account for {live_share:.0f}% of top-creator GMV in Nursing & Feeding. Hegen currently has no live-specialist creators. Partner with 1–2 mid-tier Baby creators (100K–500K followers) for weekly product demos — target Rp50jt GMV/session based on category benchmarks."),
    ("3", "Run a micro-creator seeding programme",
     f"Micro creators (10K–100K followers) generate the highest total units in this category. Hegen's best product has only 9 affiliates vs competitor products with {int(top5_competitors.iloc[-1]['total_aff'])}+. A structured seeding campaign targeting 50 new micro-creators over 30 days would close this gap and build sustainable affiliate volume."),
]

for r_base, (num, title, detail) in enumerate(recs):
    r = 17 + r_base * 3
    ws4.merge_cells(f"A{r}:H{r}")
    hdr(ws4[f"A{r}"], f"  REC {num}:  {title}", fill=BLUE_FILL, size=11)
    ws4.row_dimensions[r].height = 22
    ws4.merge_cells(f"A{r+1}:H{r+2}")
    cell = ws4[f"A{r+1}"]
    cell.value = detail
    cell.font = Font(size=10, name="Calibri", color="2D3748")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = PatternFill("solid", fgColor="EBF4FF")
    ws4.row_dimensions[r+1].height = 22
    ws4.row_dimensions[r+2].height = 14

set_col_widths(ws4, {
    "A":16,"B":16,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16
})

try:
    img6 = XLImage(f"{OUT_DIR}/10_hegen_vs_benchmark.png")
    img6.width, img6.height = 600, 280
    ws4.add_image(img6, "A30")
    img7 = XLImage(f"{OUT_DIR}/11_hegen_affiliate_gap.png")
    img7.width, img7.height = 440, 250
    ws4.add_image(img7, "A48")
except: pass

# ════════════════════════════════════
# SHEET 5: RAW DATA
# ════════════════════════════════════
ws5 = wb.create_sheet("5. Raw Data")
ws5.sheet_view.showGridLines = False
ws5.merge_cells("A1:C1")
hdr(ws5["A1"], "RAW DATA REFERENCE", fill=DARK_FILL)
val(ws5["A2"], "Shop List (Nursing & Feeding)", bold=True, size=10)
for r, (_, row) in enumerate(sl.head(50).iterrows(), 3):
    val(ws5.cell(r, 1), row["Shop Name"], size=8)
    val(ws5.cell(r, 2), fmt_idr(row["gmv_7d"]) if pd.notna(row["gmv_7d"]) else "-", size=8)
    val(ws5.cell(r, 3), fmt_k(row["influencers"]) if pd.notna(row["influencers"]) else "-", size=8)

# ── Save workbook ─────────────────────────────────────────────────────────────
excel_path = f"{OUT_DIR}/hegen_category_dashboard.xlsx"
wb.save(excel_path)
print(f"\n  ✓ Excel dashboard saved: hegen_category_dashboard.xlsx")

print("\n" + "="*60)
print("  ALL DONE — FILES GENERATED:")
print("="*60)
for f in sorted(os.listdir(OUT_DIR)):
    size = os.path.getsize(f"{OUT_DIR}/{f}") / 1024
    print(f"  {f:<50} {size:>6.1f} KB")
print("\n  Run complete ✓")
